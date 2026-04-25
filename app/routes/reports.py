"""
CIS School System - Report Card Routes
========================================
Handles:
  - Adding/editing teacher comments
  - Generating PDF report cards
  - Downloading reports (individual or ZIP)
  - Principal approval workflow
  - Email to parents
"""

import os
import zipfile
from datetime import datetime
from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, session, send_file, jsonify
)
from ..models import (
    db, Teacher, Student, Grade, Stream, Term,
    Assessment, Mark, ReportCard, get_subjects, get_split_subjects
)
from ..grading import compute_term_report, compute_class_summary, rank_students
from .auth import login_required, role_required

reports_bp = Blueprint("reports", __name__)


# ---------------------------------------------------------------------------
# HELPER — load all marks for a student across all 3 assessments
# ---------------------------------------------------------------------------

def _load_student_marks(student_id, term_id):
    """
    Returns { assessment_number: [Mark, ...] }
    """
    assessments = Assessment.query.filter_by(term_id=term_id).all()
    result = {}
    for ass in assessments:
        marks = Mark.query.filter_by(student_id=student_id, assessment_id=ass.id).all()
        result[ass.number] = marks
    return result


# ---------------------------------------------------------------------------
# COMMENTS — class teacher adds/edits the three feedback sections
# ---------------------------------------------------------------------------

@reports_bp.route("/comments/<int:stream_id>")
@login_required
def comments_list(stream_id):
    """Show all students in a stream with their comment status"""
    teacher     = Teacher.query.get(session["teacher_id"])
    stream      = Stream.query.get_or_404(stream_id)
    active_term = Term.query.filter_by(is_active=True).first()

    if not active_term:
        flash("No active term.", "warning")
        return redirect(url_for("main.dashboard"))

    # Security
    if teacher.role not in ("admin", "principal") and teacher.stream_id != stream_id:
        flash("Permission denied.", "danger")
        return redirect(url_for("main.dashboard"))

    students = (
        Student.query
        .filter_by(stream_id=stream_id, is_active=True)
        .order_by(Student.full_name)
        .all()
    )

    # Get existing report cards
    student_ids   = [s.id for s in students]
    report_cards  = ReportCard.query.filter(
        ReportCard.student_id.in_(student_ids),
        ReportCard.term_id == active_term.id
    ).all()
    rc_by_student = {rc.student_id: rc for rc in report_cards}

    return render_template(
        "reports/comments_list.html",
        stream=stream,
        students=students,
        rc_by_student=rc_by_student,
        active_term=active_term,
        teacher=teacher,
    )


@reports_bp.route("/comments/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
def edit_comments(student_id):
    """Edit the three comment sections for one student"""
    teacher     = Teacher.query.get(session["teacher_id"])
    student     = Student.query.get_or_404(student_id)
    active_term = Term.query.filter_by(is_active=True).first()

    if not active_term:
        flash("No active term.", "warning")
        return redirect(url_for("main.dashboard"))

    # Security
    if teacher.role not in ("admin", "principal") and teacher.stream_id != student.stream_id:
        flash("Permission denied.", "danger")
        return redirect(url_for("main.dashboard"))

    # Get or create report card record
    rc = ReportCard.query.filter_by(student_id=student_id, term_id=active_term.id).first()
    if not rc:
        rc = ReportCard(student_id=student_id, term_id=active_term.id, status="draft")
        db.session.add(rc)
        db.session.commit()

    if request.method == "POST":
        rc.comment_performance  = request.form.get("comment_performance", "").strip()
        rc.comment_competencies = request.form.get("comment_competencies", "").strip()
        rc.comment_values       = request.form.get("comment_values", "").strip()

        # If all three comments are filled, advance status
        if rc.comment_performance and rc.comment_competencies and rc.comment_values:
            rc.status = "pending_approval"
        else:
            rc.status = "comments_pending"

        db.session.commit()
        flash(f"Comments saved for {student.full_name}.", "success")

        # Go to next student if "Save & Next" was clicked
        if request.form.get("action") == "save_next":
            students = (
                Student.query
                .filter_by(stream_id=student.stream_id, is_active=True)
                .order_by(Student.full_name)
                .all()
            )
            ids = [s.id for s in students]
            current_idx = ids.index(student_id)
            if current_idx + 1 < len(ids):
                return redirect(url_for("reports.edit_comments", student_id=ids[current_idx + 1]))

        return redirect(url_for("reports.comments_list", stream_id=student.stream_id))

    # Load marks preview for the form
    marks_by_assessment = _load_student_marks(student_id, active_term.id)
    report_data = compute_term_report(student, active_term, marks_by_assessment)

    return render_template(
        "reports/edit_comments.html",
        student=student,
        rc=rc,
        report_data=report_data,
        teacher=teacher,
    )


# ---------------------------------------------------------------------------
# GENERATE PDF — for one student
# ---------------------------------------------------------------------------

@reports_bp.route("/generate/<int:student_id>")
@login_required
def generate_one(student_id):
    """Generate a PDF report card for one student"""
    from ..pdf_generator import generate_report_pdf

    teacher     = Teacher.query.get(session["teacher_id"])
    student     = Student.query.get_or_404(student_id)
    active_term = Term.query.filter_by(is_active=True).first()

    if not active_term:
        flash("No active term.", "warning")
        return redirect(url_for("main.dashboard"))

    rc = ReportCard.query.filter_by(student_id=student_id, term_id=active_term.id).first()
    if not rc:
        flash("Please add comments before generating the report.", "warning")
        return redirect(url_for("reports.edit_comments", student_id=student_id))

    marks_by_assessment = _load_student_marks(student_id, active_term.id)
    report_data         = compute_term_report(student, active_term, marks_by_assessment)

    # Add comments and class teacher name to report data
    report_data["comments"] = {
        "performance":   rc.comment_performance  or "",
        "competencies":  rc.comment_competencies or "",
        "values":        rc.comment_values        or "",
    }
    if student.stream and student.stream.teachers:
        report_data["class_teacher"] = student.stream.teachers[0].full_name
    else:
        report_data["class_teacher"] = ""

    pdf_path = generate_report_pdf(report_data, student)

    rc.pdf_path     = pdf_path
    rc.generated_at = datetime.utcnow()
    db.session.commit()

    return send_file(pdf_path, as_attachment=True, download_name=f"{student.full_name}_Report.pdf")


# ---------------------------------------------------------------------------
# GENERATE ZIP — all reports for a stream
# ---------------------------------------------------------------------------

@reports_bp.route("/generate/stream/<int:stream_id>")
@login_required
@role_required("admin", "principal")
def generate_stream_zip(stream_id):
    """Generate PDFs for all students in a stream and zip them for download"""
    from ..pdf_generator import generate_report_pdf
    import tempfile

    stream      = Stream.query.get_or_404(stream_id)
    active_term = Term.query.filter_by(is_active=True).first()

    students = (
        Student.query
        .filter_by(stream_id=stream_id, is_active=True)
        .order_by(Student.full_name)
        .all()
    )

    if not students:
        flash("No students found in this stream.", "warning")
        return redirect(url_for("admin.index"))

    from flask import current_app
    reports_folder = current_app.config["REPORTS_FOLDER"]
    os.makedirs(reports_folder, exist_ok=True)

    zip_filename = f"Reports_{stream.grade.name}_{stream.name}_Term{active_term.term_number}_{active_term.academic_year.year}.zip"
    zip_path     = os.path.join(reports_folder, zip_filename)

    with zipfile.ZipFile(zip_path, "w") as zf:
        for student in students:
            rc = ReportCard.query.filter_by(student_id=student.id, term_id=active_term.id).first()

            marks_by_assessment = _load_student_marks(student.id, active_term.id)
            report_data         = compute_term_report(student, active_term, marks_by_assessment)

            report_data["comments"] = {
                "performance":  rc.comment_performance  if rc else "",
                "competencies": rc.comment_competencies if rc else "",
                "values":       rc.comment_values        if rc else "",
            }
            if student.stream and student.stream.teachers:
                report_data["class_teacher"] = student.stream.teachers[0].full_name
            else:
                report_data["class_teacher"] = ""

            pdf_path = generate_report_pdf(report_data, student)
            zf.write(pdf_path, arcname=f"{student.full_name}_Report.pdf")

            if rc:
                rc.pdf_path     = pdf_path
                rc.generated_at = datetime.utcnow()

    db.session.commit()
    return send_file(zip_path, as_attachment=True, download_name=zip_filename)


# ---------------------------------------------------------------------------
# APPROVAL WORKFLOW
# ---------------------------------------------------------------------------

@reports_bp.route("/approve/<int:report_id>", methods=["POST"])
@login_required
@role_required("principal", "admin")
def approve_report(report_id):
    """Principal approves a report card — marks it ready for printing/emailing"""
    rc = ReportCard.query.get_or_404(report_id)
    rc.status      = "approved"
    rc.approved_by = session["teacher_id"]
    rc.approved_at = datetime.utcnow()
    db.session.commit()
    flash(f"Report approved for {rc.student.full_name}.", "success")
    return redirect(request.referrer or url_for("admin.index"))


# ---------------------------------------------------------------------------
# OVERVIEW — all streams completion status
# ---------------------------------------------------------------------------

@reports_bp.route("/overview")
@login_required
@role_required("admin", "principal")
def overview():
    """Show report generation progress for all streams"""
    active_term = Term.query.filter_by(is_active=True).first()
    grades      = Grade.query.order_by(Grade.sort_order).all()

    progress = {}
    if active_term:
        for grade in grades:
            for stream in grade.streams:
                total   = Student.query.filter_by(stream_id=stream.id, is_active=True).count()
                student_ids = [
                    s.id for s in Student.query.filter_by(stream_id=stream.id, is_active=True).all()
                ]
                with_comments = ReportCard.query.filter(
                    ReportCard.student_id.in_(student_ids),
                    ReportCard.term_id == active_term.id,
                    ReportCard.comment_performance != None,
                ).count()
                approved = ReportCard.query.filter(
                    ReportCard.student_id.in_(student_ids),
                    ReportCard.term_id == active_term.id,
                    ReportCard.status == "approved",
                ).count()

                progress[stream.id] = {
                    "stream":        stream,
                    "total":         total,
                    "with_comments": with_comments,
                    "approved":      approved,
                }

    return render_template(
        "reports/overview.html",
        grades=grades,
        active_term=active_term,
        progress=progress,
    )


@reports_bp.route("/download/all-pdfs")
@login_required
@role_required("admin", "principal")
def download_all_pdfs():
    """Generate and ZIP all report cards for every student in the school."""
    import zipfile
    from datetime import datetime
    from ..pdf_generator import generate_report_pdf
    from flask import current_app
    from ..models import Grade, Stream

    active_term = Term.query.filter_by(is_active=True).first()
    if not active_term:
        flash("No active term.", "warning")
        return redirect(url_for("reports.overview"))

    reports_folder = current_app.config["REPORTS_FOLDER"]
    os.makedirs(reports_folder, exist_ok=True)

    zip_filename = f"CIS_ALL_Reports_Term{active_term.term_number}_{active_term.academic_year.year}.zip"
    zip_path = os.path.join(reports_folder, zip_filename)

    grades = Grade.query.order_by(Grade.sort_order).all()
    total = 0

    with zipfile.ZipFile(zip_path, "w") as zf:
        for grade in grades:
            for stream in grade.streams:
                students = (
                    Student.query
                    .filter_by(stream_id=stream.id, is_active=True)
                    .order_by(Student.full_name)
                    .all()
                )
                for student in students:
                    rc = ReportCard.query.filter_by(
                        student_id=student.id, term_id=active_term.id
                    ).first()

                    marks_by_assessment = _load_student_marks(student.id, active_term.id)
                    report_data = compute_term_report(student, active_term, marks_by_assessment)

                    report_data["comments"] = {
                        "performance":   rc.comment_performance  if rc else "",
                        "competencies":  rc.comment_competencies if rc else "",
                        "values":        rc.comment_values       if rc else "",
                        "general":       rc.general_comment      if rc else "",
                    }
                    if student.stream and student.stream.teachers:
                        report_data["class_teacher"] = student.stream.teachers[0].full_name
                    else:
                        report_data["class_teacher"] = ""

                    try:
                        pdf_path = generate_report_pdf(report_data, student)
                        # Organise into folders by grade
                        folder = f"{grade.name}/{stream.name}"
                        zf.write(pdf_path, arcname=f"{folder}/{student.full_name}_Report.pdf")
                        total += 1
                    except Exception as e:
                        print(f"  ⚠ Skipped {student.full_name}: {e}")

    return send_file(zip_path, as_attachment=True, download_name=zip_filename)


@reports_bp.route("/download/marks-excel")
@login_required
@role_required("admin", "principal")
def download_marks_excel():
    """Export all marks for all students to a single Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import current_app
    from ..models import Grade, get_subjects, get_split_subjects

    active_term = Term.query.filter_by(is_active=True).first()
    if not active_term:
        flash("No active term.", "warning")
        return redirect(url_for("reports.overview"))

    reports_folder = current_app.config["REPORTS_FOLDER"]
    os.makedirs(reports_folder, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Styles
    hdr_font    = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill    = PatternFill("solid", fgColor="002147")
    subhdr_fill = PatternFill("solid", fgColor="C8962D")
    subhdr_font = Font(bold=True, color="FFFFFF", size=9)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    assessments = Assessment.query.filter_by(term_id=active_term.id).order_by(Assessment.number).all()
    grades      = Grade.query.order_by(Grade.sort_order).all()

    for grade in grades:
        subjects     = get_subjects(grade.name)
        split        = get_split_subjects(grade.name)
        sheet_name   = grade.name[:31]
        ws           = wb.create_sheet(title=sheet_name)

        # ── Build column headers ──────────────────────────────────
        # Row 1: grade + term info
        ws.merge_cells("A1:B1")
        ws["A1"] = f"{grade.name}  —  Term {active_term.term_number}, {active_term.academic_year.year}"
        ws["A1"].font = Font(bold=True, size=12, color="002147")

        # Row 2: section headers
        row2 = ["#", "Name", "Stream"]
        for ass in assessments:
            for subj in subjects:
                if subj in split:
                    p1l, _, p2l, _ = split[subj]
                    row2 += [f"{subj} ({p1l})", f"{subj} ({p2l})", f"{subj} %"]
                else:
                    row2.append(subj)
            row2 += ["Total", "Average", "PL"]
        row2.append("Comment")

        for col_idx, val in enumerate(row2, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font      = hdr_font if col_idx <= 3 else subhdr_font
            cell.fill      = hdr_fill if col_idx <= 3 else subhdr_fill
            cell.alignment = center
            cell.border    = thin

        # ── Data rows ─────────────────────────────────────────────
        all_students = (
            Student.query
            .filter_by(grade_id=grade.id, is_active=True)
            .order_by(Student.stream_id, Student.full_name)
            .all()
        )

        for row_num, student in enumerate(all_students, 3):
            rc   = ReportCard.query.filter_by(student_id=student.id, term_id=active_term.id).first()
            cols = [row_num - 2, student.full_name, student.stream.name if student.stream else ""]

            for ass in assessments:
                marks_dict = {
                    m.subject: m for m in
                    Mark.query.filter_by(student_id=student.id, assessment_id=ass.id).all()
                }
                total = 0.0
                count = 0
                for subj in subjects:
                    m = marks_dict.get(subj)
                    if subj in split:
                        p1 = m.paper1_score   if m else None
                        p2 = m.paper2_score   if m else None
                        pc = m.combined_score if m else None
                        cols += [p1 or "", p2 or "", pc or ""]
                        if pc:
                            total += pc; count += 1
                    else:
                        sc = m.score if m else None
                        cols.append(sc if sc is not None else "")
                        if sc is not None:
                            total += sc; count += 1

                avg = round(total / count, 2) if count else ""
                pl  = marks_dict[subjects[0]].grade_code if subjects and subjects[0] in marks_dict and marks_dict[subjects[0]].grade_code else ""
                cols += [round(total, 2) if count else "", avg, pl]

            comment = ""
            if rc:
                comment = rc.comment_performance or rc.general_comment or ""
            cols.append(comment)

            for col_idx, val in enumerate(cols, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = left if col_idx <= 3 else center
                cell.border    = thin
                if row_num % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="E8F0FA")

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

    filename = f"CIS_Marks_Term{active_term.term_number}_{active_term.academic_year.year}.xlsx"
    filepath = os.path.join(reports_folder, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)
